#!/usr/bin/env python3
"""Build neurotech_dataset.xlsx from enriched CSV."""
import csv
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule

ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / 'data' / 'processed' / 'neurotech_enriched.csv'
OUT_PATH = ROOT / 'local' / 'neurotech_dataset.xlsx'
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# ----- Styles -----
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', start_color='1F4E79')
BODY_FONT = Font(name='Arial', size=10)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F4E79')
NOTE_FONT = Font(name='Arial', italic=True, size=9, color='595959')
THIN = Side(border_style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws, max_width=40):
    for col in ws.columns:
        maxlen = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            l = len(str(cell.value))
            if l > maxlen:
                maxlen = l
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), max_width)


# ----- Load data -----
rows = []
with CSV_PATH.open() as f:
    for r in csv.DictReader(f):
        rows.append(r)

# ----- Build workbook -----
wb = Workbook()

# --------- Sheet 1: README ---------
ws = wb.active
ws.title = 'README'
ws['A1'] = 'Neurotech Company Dataset'
ws['A1'].font = TITLE_FONT
ws['A2'] = f'Built: {datetime.now().strftime("%Y-%m-%d")}  |  Source: app.reccy.dev (reccy neuro)  |  Records: {len(rows)}'
ws['A2'].font = NOTE_FONT

ws['A4'] = 'About this dataset'
ws['A4'].font = Font(name='Arial', bold=True, size=12)

notes = [
    '',
    'This dataset was built as an independent alternative to the Centre for Future Generations (CFG)',
    '"Neurotech Consumer Market Atlas" (271 companies, Jan 2025 cutoff). It uses the reccy neuro company',
    'board (app.reccy.dev) as the base universe — 395 companies total, scraped directly from the app.',
    '',
    'Why build a new dataset (vs. reuse CFG):',
    '  • CFG has ~31% fewer companies (271 vs 395)',
    '  • CFG excludes adjacent sectors (mental health wearables, neurofeedback, eye-tracking, pharma)',
    '  • CFG underrepresents Asia (only 11%)',
    '  • CFG data cutoff is Jan 2025 — already ~15 months stale',
    '  • CFG financial data is only ~62% complete',
    '  • CFG has only 3 historical phases, not yearly/half-year granularity',
    '  • Independent dataset enables cross-validation',
    '',
    'Sheets in this workbook',
    '  • companies         — full enriched dataset (one row per company)',
    '  • timeline_year     — founding counts by year, and by modality/region',
    '  • timeline_decade   — births aggregated by decade',
    '  • geo               — country & region breakdown',
    '  • funding           — last-round-stage distribution',
    '  • sectors           — industry-tag frequency & cross-tabs',
    '  • methodology       — how this was built, limitations, confidence',
    '',
    'Critical caveats (please read before using)',
    '  • Founding year coverage is PARTIAL. Only ~48% of rows have a founding year populated from',
    '    training knowledge. The rest need additional web enrichment (LinkedIn, Crunchbase, company sites).',
    '  • Founding year confidence tiers:  H = high, M = medium, L = rough estimate.',
    '  • Industries are classifier tags from reccy, not curated. Modality/application/invasiveness are',
    '    derived heuristically — treat as first-pass labels, not ground truth.',
    '  • Funding data is stage-only (Series A, Series B, Grant, etc.) — no dollar amounts.',
    '  • Some name strings are messy (e.g. "FRENZ™ by Earable™") — kept verbatim for traceability.',
    '',
    'How to improve this dataset (next steps)',
    '  1. Web-search the ~205 companies without founding year, in batches of ~10 per query',
    '  2. Pull total-raised USD from Crunchbase/Dealroom where available',
    '  3. Manually review the heuristic modality/invasiveness labels',
    '  4. Cross-check names against CFG list to see overlap and unique entries',
    '  5. Add status tracking (active / acquired / shutdown) via web search',
]
for i, line in enumerate(notes, start=5):
    ws.cell(row=i, column=1, value=line).font = BODY_FONT
ws.column_dimensions['A'].width = 110

# --------- Sheet 2: companies ---------
ws = wb.create_sheet('companies')
headers = [
    'name', 'website', 'country', 'region', 'location', 'founding_year',
    'founding_confidence', 'decade', 'primary_modality', 'application',
    'invasiveness', 'last_funding_stage', 'headcount_bucket', 'live_jobs',
    'industries',
]
for c, h in enumerate(headers, start=1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, 1, len(headers))
ws.freeze_panes = 'A2'

for ri, r in enumerate(rows, start=2):
    ws.cell(row=ri, column=1, value=r['name'])
    ws.cell(row=ri, column=2, value=r['website'])
    ws.cell(row=ri, column=3, value=r['country'])
    ws.cell(row=ri, column=4, value=r['region'])
    ws.cell(row=ri, column=5, value=r['location'])
    yr = r['founding_year']
    ws.cell(row=ri, column=6, value=int(yr) if yr else None)
    ws.cell(row=ri, column=7, value=r['founding_confidence'])
    ws.cell(row=ri, column=8, value=r['decade'])
    ws.cell(row=ri, column=9, value=r['primary_modality'])
    ws.cell(row=ri, column=10, value=r['application'])
    ws.cell(row=ri, column=11, value=r['invasiveness'])
    ws.cell(row=ri, column=12, value=r['last_funding_stage'])
    ws.cell(row=ri, column=13, value=r['headcount_bucket'])
    lj = r['live_jobs']
    ws.cell(row=ri, column=14, value=int(lj) if lj and lj.isdigit() else 0)
    ws.cell(row=ri, column=15, value=r['industries'])

last_row = len(rows) + 1
# Format founding_year as plain number (no thousand sep)
for c in range(2, last_row + 1):
    ws.cell(row=c, column=6).number_format = '0'

# Add filter
ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{last_row}'
autosize(ws, max_width=45)

# --------- Sheet 3: timeline_year ---------
ws = wb.create_sheet('timeline_year')
ws['A1'] = 'Founding-year births (companies founded per year)'
ws['A1'].font = TITLE_FONT
ws['A2'] = 'NOTE: Only ~48% of companies have a founding year populated. Counts are therefore a lower bound.'
ws['A2'].font = NOTE_FONT

# Determine year range
years = sorted(set(int(r['founding_year']) for r in rows if r['founding_year']))
year_min, year_max = min(years), max(years)
year_list = list(range(year_min, year_max + 1))

ws['A4'] = 'Year'
ws['B4'] = 'Births (total)'
ws['C4'] = 'North America'
ws['D4'] = 'Europe'
ws['E4'] = 'Asia'
ws['F4'] = 'MENA'
ws['G4'] = 'Oceania'
ws['H4'] = 'Other'
ws['I4'] = 'Invasive'
ws['J4'] = 'Non-invasive'
style_header(ws, 4, 1, 10)

comp_last = len(rows) + 1  # last data row in companies sheet
for i, yr in enumerate(year_list):
    r = 5 + i
    ws.cell(row=r, column=1, value=yr).number_format = '0'
    ws.cell(row=r, column=2, value=f'=COUNTIF(companies!F2:F{comp_last},A{r})')
    for col_letter, region in zip(['C', 'D', 'E', 'F', 'G', 'H'],
                                   ['North America', 'Europe', 'Asia', 'MENA', 'Oceania', 'Other']):
        ws.cell(row=r, column=ord(col_letter) - 64,
                value=f'=COUNTIFS(companies!F2:F{comp_last},A{r},companies!D2:D{comp_last},"{region}")')
    ws.cell(row=r, column=9, value=f'=COUNTIFS(companies!F2:F{comp_last},A{r},companies!K2:K{comp_last},"Invasive")')
    ws.cell(row=r, column=10, value=f'=COUNTIFS(companies!F2:F{comp_last},A{r},companies!K2:K{comp_last},"Non-invasive")')

total_row = 5 + len(year_list)
ws.cell(row=total_row, column=1, value='Total').font = Font(bold=True)
for col in range(2, 11):
    letter = get_column_letter(col)
    ws.cell(row=total_row, column=col, value=f'=SUM({letter}5:{letter}{total_row-1})').font = Font(bold=True)

# Add chart: births over time
chart = BarChart()
chart.type = 'col'
chart.style = 11
chart.title = 'Neurotech company births per year (known founding years)'
chart.y_axis.title = 'Companies founded'
chart.x_axis.title = 'Year'
data = Reference(ws, min_col=2, min_row=4, max_col=2, max_row=total_row - 1)
cats = Reference(ws, min_col=1, min_row=5, max_row=total_row - 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 22
chart.height = 10
ws.add_chart(chart, 'L4')

autosize(ws, max_width=22)

# --------- Sheet 4: timeline_decade ---------
ws = wb.create_sheet('timeline_decade')
ws['A1'] = 'Founding births by decade'
ws['A1'].font = TITLE_FONT
decades = ['1940s', '1950s', '1960s', '1970s', '1980s', '1990s', '2000s', '2010s', '2020s']
ws['A3'] = 'Decade'
ws['B3'] = 'Births'
ws['C3'] = '% of known'
style_header(ws, 3, 1, 3)
for i, d in enumerate(decades):
    r = 4 + i
    ws.cell(row=r, column=1, value=d)
    ws.cell(row=r, column=2, value=f'=COUNTIF(companies!H2:H{comp_last},A{r})')
    ws.cell(row=r, column=3, value=f'=IF(SUM($B$4:$B$12)=0,0,B{r}/SUM($B$4:$B$12))')
    ws.cell(row=r, column=3).number_format = '0.0%'

chart = BarChart()
chart.type = 'col'
chart.style = 12
chart.title = 'Births by decade'
chart.y_axis.title = 'Companies'
data = Reference(ws, min_col=2, min_row=3, max_col=2, max_row=12)
cats = Reference(ws, min_col=1, min_row=4, max_row=12)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 18
chart.height = 9
ws.add_chart(chart, 'E3')
autosize(ws)

# --------- Sheet 5: geo ---------
ws = wb.create_sheet('geo')
ws['A1'] = 'Geographic distribution'
ws['A1'].font = TITLE_FONT

ws['A3'] = 'Country'
ws['B3'] = 'Companies'
ws['C3'] = '% of total'
style_header(ws, 3, 1, 3)
from collections import Counter
cc = Counter(r['country'] for r in rows if r['country'])
countries_sorted = cc.most_common()
for i, (country, count) in enumerate(countries_sorted):
    r = 4 + i
    ws.cell(row=r, column=1, value=country)
    ws.cell(row=r, column=2, value=f'=COUNTIF(companies!C2:C{comp_last},A{r})')
    ws.cell(row=r, column=3, value=f'=B{r}/COUNTA(companies!C2:C{comp_last})')
    ws.cell(row=r, column=3).number_format = '0.0%'

last_geo = 4 + len(countries_sorted) - 1

# Regional summary
rc_start = 3
rc_col = 5
ws.cell(row=3, column=rc_col, value='Region')
ws.cell(row=3, column=rc_col + 1, value='Companies')
style_header(ws, 3, rc_col, rc_col + 1)
rc = Counter(r['region'] for r in rows if r['region'])
for i, (region, count) in enumerate(rc.most_common()):
    r = 4 + i
    ws.cell(row=r, column=rc_col, value=region)
    ws.cell(row=r, column=rc_col + 1, value=f'=COUNTIF(companies!D2:D{comp_last},E{r})')

# Chart: top 15 countries
chart = BarChart()
chart.type = 'bar'
chart.style = 13
chart.title = 'Top countries by company count'
chart.y_axis.title = 'Country'
top_n = min(15, len(countries_sorted))
data = Reference(ws, min_col=2, min_row=3, max_col=2, max_row=3 + top_n)
cats = Reference(ws, min_col=1, min_row=4, max_row=3 + top_n)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 18
chart.height = 12
ws.add_chart(chart, 'H3')
autosize(ws)

# --------- Sheet 6: funding ---------
ws = wb.create_sheet('funding')
ws['A1'] = 'Last funding stage distribution'
ws['A1'].font = TITLE_FONT

ws['A3'] = 'Last round stage'
ws['B3'] = 'Companies'
ws['C3'] = '% of total'
style_header(ws, 3, 1, 3)
fc = Counter((r['last_funding_stage'] or '(unknown)') for r in rows)
for i, (stage, count) in enumerate(fc.most_common()):
    r = 4 + i
    ws.cell(row=r, column=1, value=stage)
    ws.cell(row=r, column=2, value=f'=COUNTIF(companies!L2:L{comp_last},IF(A{r}="(unknown)","",A{r}))')
    ws.cell(row=r, column=3, value=f'=B{r}/{len(rows)}')
    ws.cell(row=r, column=3).number_format = '0.0%'

chart = BarChart()
chart.type = 'bar'
chart.style = 14
chart.title = 'Companies by last funding stage'
top_n = min(20, len(fc))
data = Reference(ws, min_col=2, min_row=3, max_col=2, max_row=3 + top_n)
cats = Reference(ws, min_col=1, min_row=4, max_row=3 + top_n)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 18
chart.height = 12
ws.add_chart(chart, 'E3')
autosize(ws)

# --------- Sheet 7: sectors ---------
ws = wb.create_sheet('sectors')
ws['A1'] = 'Primary modality, application & invasiveness breakdown'
ws['A1'].font = TITLE_FONT

ws['A3'] = 'Primary modality'
ws['B3'] = 'Companies'
style_header(ws, 3, 1, 2)
mc = Counter(r['primary_modality'] for r in rows)
for i, (k, v) in enumerate(mc.most_common()):
    r = 4 + i
    ws.cell(row=r, column=1, value=k)
    ws.cell(row=r, column=2, value=f'=COUNTIF(companies!I2:I{comp_last},A{r})')

# Application
col_off = 4
ws.cell(row=3, column=col_off, value='Application')
ws.cell(row=3, column=col_off + 1, value='Companies')
style_header(ws, 3, col_off, col_off + 1)
ac = Counter(r['application'] for r in rows)
for i, (k, v) in enumerate(ac.most_common()):
    r = 4 + i
    ws.cell(row=r, column=col_off, value=k)
    ws.cell(row=r, column=col_off + 1, value=f'=COUNTIF(companies!J2:J{comp_last},{get_column_letter(col_off)}{r})')

# Invasiveness
col_off = 7
ws.cell(row=3, column=col_off, value='Invasiveness')
ws.cell(row=3, column=col_off + 1, value='Companies')
style_header(ws, 3, col_off, col_off + 1)
ic = Counter(r['invasiveness'] for r in rows)
for i, (k, v) in enumerate(ic.most_common()):
    r = 4 + i
    ws.cell(row=r, column=col_off, value=k)
    ws.cell(row=r, column=col_off + 1, value=f'=COUNTIF(companies!K2:K{comp_last},{get_column_letter(col_off)}{r})')

autosize(ws, max_width=25)

# --------- Sheet 8: methodology ---------
ws = wb.create_sheet('methodology')
ws['A1'] = 'Methodology & limitations'
ws['A1'].font = TITLE_FONT
meth = [
    '',
    '1. DATA COLLECTION',
    '   Base universe: reccy.dev "Neurotech Company Board" (app.reccy.dev/companies)',
    '   Accessed: 2026-04-09 via Claude-in-Chrome browser automation',
    '   Extraction: Direct extraction from the React fiber (not DOM scraping) — all 395 companies',
    '   from the internal `companies` prop captured in a single pass.',
    '   Fields pulled from reccy: id, name, industries[], headcount, lastFundingType, website,',
    '   location, jobCount. (logoUrl stripped for privacy.)',
    '',
    '2. DERIVED FIELDS',
    '   country       — parsed from location string via country/state mapping',
    '   region        — North America / Europe / Asia / MENA / Oceania / LATAM / Other',
    '   primary_modality — heuristic: Wearable / Hardware / Software-AI / Research Tools /',
    '                      Therapeutics / Diagnostics / Medical Device / Other',
    '   application   — Medical-Therapeutic / Medical-Diagnostic / Consumer-Wellness /',
    '                   Research Tools / Software-Analytics / Medical-Device / Other',
    '   invasiveness  — Invasive / Non-invasive / Mixed — known invasive BCI companies are',
    '                   hardcoded; everything else defaults to non-invasive (imperfect heuristic!)',
    '',
    '3. FOUNDING YEAR',
    '   Source: training-time knowledge of the author model for ~50% of companies',
    '   Confidence tiers:  H (public info, highly confident) — well-known/public companies',
    '                      M (medium) — reasonable estimate from training data',
    '                      L (low)    — rough guess, should be treated as unverified',
    '   COVERAGE: ~48% of companies have a populated founding year.',
    '   The remaining ~52% need web-enrichment via LinkedIn/Crunchbase/company websites.',
    '',
    '4. KNOWN LIMITATIONS',
    '   • Two companies were lost during row deduplication (393 of 395 retained)',
    '   • Industries are reccy classifier tags, not curated — labels overlap and are inconsistent',
    '   • Headcount is a bucket, not a count',
    '   • Funding is stage-only — no dollar amounts or round dates',
    '   • No status tracking (active / acquired / shutdown)',
    '   • Modality/application/invasiveness are rule-based heuristics, not human-labeled',
    '   • Region classification treats Israel/UAE as "MENA" — may not match other taxonomies',
    '',
    '5. WHAT WOULD MAKE THIS DATASET 10x BETTER',
    '   • Complete founding years via web search (realistic target: 85-90%)',
    '   • Total raised USD from Crunchbase/Dealroom',
    '   • Founder backgrounds (academic vs industry)',
    '   • Clinical trials count (from ClinicalTrials.gov) for medical companies',
    '   • Patent counts (from Google Patents)',
    '   • FDA clearance / CE mark timeline',
    '   • Cross-validation with CFG Neurotech Market Atlas (271 companies)',
    '   • Death/acquisition tracking to compute survival curves',
]
for i, line in enumerate(meth, start=2):
    ws.cell(row=i, column=1, value=line).font = BODY_FONT
ws.column_dimensions['A'].width = 110

# Reorder: put README first, then companies
wb.move_sheet('README', offset=-wb.sheetnames.index('README'))

wb.save(OUT_PATH)
print(f'Saved: {OUT_PATH}')
